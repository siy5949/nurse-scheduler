import streamlit as st
import pandas as pd
import random
import calendar
import holidays
import os
import io
import time
import copy
import glob
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ==========================================
# 1. 설정 및 상수
# ==========================================
st.set_page_config(layout="wide", page_title="5병동 근무표 시스템")
kr_holidays = holidays.KR()
DB_FILE = "staff_db.csv"
SCHEDULE_DIR = "saved_schedules"

if not os.path.exists(SCHEDULE_DIR):
    os.makedirs(SCHEDULE_DIR)

# 색상
COLOR_REQ_OFF = "FFFF00"  # 쨍한 노랑
COLOR_DE = "FF8C00"       # 쨍한 주황 (Dark Orange)
COLOR_N = "F2F2F2"        # 연한 회색
COLOR_WEEKEND_BG = "FFD9D9"
COLOR_WHITE = "FFFFFF"
COLOR_ERROR = "FF0000"    # 오류

ROLE_ORDER = {"HN": 0, "RN": 1, "AN": 2}
ROLES = ["RN", "HN", "AN"]
MAX_N_LIMIT = 10    
MAX_N_EXTENDED = 11 
MIN_OFF_LIMIT = 6   

# ==========================================
# 2. 데이터 관리
# ==========================================
def load_staff_data():
    if os.path.exists(DB_FILE):
        try: return pd.read_csv(DB_FILE).to_dict('records')
        except: return []
    return [
        {"name": "김유진", "role": "HN", "req_off": "", "fixed_work": "", "annual_leave": 0},
        {"name": "이다슬", "role": "RN", "req_off": "", "fixed_work": "", "annual_leave": 0},
        {"name": "장완수", "role": "RN", "req_off": "", "fixed_work": "", "annual_leave": 0},
        {"name": "박지유", "role": "RN", "req_off": "", "fixed_work": "", "annual_leave": 0},
        {"name": "황경순", "role": "AN", "req_off": "", "fixed_work": "", "annual_leave": 0}
    ]

def save_staff_data(staff_list):
    pd.DataFrame(staff_list).to_csv(DB_FILE, index=False)

def save_schedule_file(df, year, month):
    filename = f"{SCHEDULE_DIR}/schedule_{year}_{month}.csv"
    df.to_csv(filename, encoding='utf-8-sig')

def load_schedule_file(year, month):
    filename = f"{SCHEDULE_DIR}/schedule_{year}_{month}.csv"
    if os.path.exists(filename):
        try:
            df = pd.read_csv(filename, index_col=0).fillna("OFF").astype(str)
            new_cols = []
            for c in df.columns:
                if c.isdigit(): new_cols.append(int(c))
                else: new_cols.append(c)
            df.columns = new_cols
            return df
        except: return None
    return None

def load_schedule_file_path(path):
    if os.path.exists(path):
        try:
            df = pd.read_csv(path, index_col=0).fillna("OFF").astype(str)
            new_cols = []
            for c in df.columns:
                if c.isdigit(): new_cols.append(int(c))
                else: new_cols.append(c)
            df.columns = new_cols
            return df
        except: return None
    return None

def delete_schedule_file(year, month):
    filename = f"{SCHEDULE_DIR}/schedule_{year}_{month}.csv"
    if os.path.exists(filename):
        os.remove(filename)

# 세션 상태
if "staff_list" not in st.session_state:
    st.session_state.staff_list = load_staff_data()
if "edit_index" not in st.session_state:
    st.session_state.edit_index = None
if "backup_staff_list" not in st.session_state:
    st.session_state.backup_staff_list = None
if "prev_df_res" not in st.session_state:
    st.session_state.prev_df_res = None
if "prev_req_map" not in st.session_state:
    st.session_state.prev_req_map = None
if "view_mode" not in st.session_state:
    st.session_state.view_mode = "create"
if "history_ym" not in st.session_state:
    st.session_state.history_ym = (0, 0)
if "edit_mode" not in st.session_state:
    st.session_state.edit_mode = False

# ★ 2월 근무표 데이터 강제 초기화
def init_feb_schedule():
    target_file = f"{SCHEDULE_DIR}/schedule_2026_2.csv"
    feb_data = {
        "김유진": ["D","D","D","D","D","E","OFF","OFF","D","D","D","D","D","OFF","E","E","E","OFF","D","D","DE","OFF","OFF","D","D","D","D","D"],
        "이다슬": ["E","E","OFF","N","N","N","OFF","DE","E","E","OFF","E","E","E","N","N","N","OFF","OFF","OFF","N","N","N","OFF","OFF","OFF","E","E"],
        "장완수": ["N","N","N","OFF","OFF","D","DE","OFF","OFF","N","N","OFF","OFF","D","D","D","OFF","N","N","N","OFF","OFF","D","E","E","E","N","N"], 
        "박지유": ["OFF","OFF","E","E","E","OFF","N","N","N","OFF","E","N","N","N","OFF","OFF","D","DE","E","E","OFF","DE","E","N","N","N","OFF","OFF"],
        "황경순": ["OFF","M","M","M","M","M","OFF","OFF","M","M","M","M","M","OFF","OFF","OFF","OFF","OFF","M","M","OFF","OFF","M","M","M","M","M","OFF"]
    }
    cols = range(1, 29)
    df = pd.DataFrame("", index=feb_data.keys(), columns=cols)
    for name, shifts in feb_data.items():
        for i, s in enumerate(shifts):
            if i < 28: df.loc[name, i+1] = s
    df.to_csv(target_file, encoding='utf-8-sig')

if "feb_initialized" not in st.session_state:
    init_feb_schedule()
    st.session_state.feb_initialized = True

def get_holidays_in_month(year, month):
    last = calendar.monthrange(year, month)[1]
    return {d for d in range(1, last+1) if date(year, month, d) in kr_holidays}

def is_holiday_or_weekend(year, month, day, hol_set):
    d = date(year, month, day)
    return d.weekday() >= 5 or day in hol_set

# ==========================================
# 3. 제약 조건
# ==========================================
def check_possibility(df, name, day, shift, req_off_list, allowed_shifts=None, strict=True):
    last_day = len(df.columns)
    if day > last_day or day < 1: return False

    if allowed_shifts is not None:
        if shift not in allowed_shifts: return False

    if df.loc[name, day] != "": return False
    if day in req_off_list: return False
    
    if day > 1:
        prev = df.loc[name, day-1]
        if prev == "N" and shift != "OFF": return False
        if prev == "E" and shift in ["D", "DE"]: return False
        if prev == "DE" and shift in ["D", "DE"]: return False
    if day < last_day:
        next_s = df.loc[name, day+1]
        if shift == "N" and next_s not in ["", "OFF"]: return False

    if shift != "OFF":
        limit = 6
        backward_streak = 0
        for i in range(1, limit + 1):
            if day-i < 1: break
            if df.loc[name, day-i] not in ["OFF", ""]: backward_streak += 1
            else: break
        forward_streak = 0
        for i in range(1, limit + 1):
            if day+i > last_day: break
            if df.loc[name, day+i] not in ["OFF", ""]: forward_streak += 1
            else: break
        if (backward_streak + 1 + forward_streak) > limit: return False
        
    return True

# ==========================================
# 4. 스케줄링 로직
# ==========================================
def attempt_schedule(year, month, staff_data, hol_set, last_day):
    names = [s['name'] for s in staff_data]
    df = pd.DataFrame("", index=names, columns=range(1, last_day + 1))
    
    rn_list = [s['name'] for s in staff_data if s['role'] == "RN"]
    hn_list = [s['name'] for s in staff_data if s['role'] == "HN"]
    an_list = [s['name'] for s in staff_data if s['role'] == "AN"]
    target_staff = hn_list + rn_list
    roles = {s['name']: s['role'] for s in staff_data}
    senior_rns = rn_list[:2] if len(rn_list) >= 2 else rn_list

    work_counts = {n: 0 for n in names}
    n_counts = {n: 0 for n in names}
    de_counts = {n: 0 for n in names}
    hn_e_counts = {n: 0 for n in names}
    
    req_off_map = {}
    fixed_work_map = {} 
    
    for s in staff_data:
        nm = s['name']
        off_str = str(s['req_off'])
        reqs = []
        if off_str and off_str.lower() != "nan":
            reqs = [int(x.strip()) for x in off_str.split(',') if x.strip().isdigit()]
        req_off_map[nm] = reqs
        
        fixed_map = {}
        fix_str = str(s.get('fixed_work', ''))
        if fix_str and fix_str.lower() != "nan":
            items = fix_str.split(',')
            for item in items:
                if '=' in item:
                    d_part, s_part = item.split('=')
                    if d_part.strip().isdigit():
                        day_num = int(d_part.strip())
                        shifts = [code.strip().upper() for code in s_part.split('/')]
                        fixed_map[day_num] = shifts
        fixed_work_map[nm] = fixed_map

    for nm in an_list:
        for d in range(1, last_day + 1):
            if d in req_off_map.get(nm, []): df.loc[nm, d] = "OFF"
            elif is_holiday_or_weekend(year, month, d, hol_set): df.loc[nm, d] = "OFF"
            else: df.loc[nm, d] = "M"
            
    for nm in names:
        if nm in fixed_work_map:
            for d, shifts in fixed_work_map[nm].items():
                if 1 <= d <= last_day and len(shifts) == 1:
                    shift = shifts[0]
                    df.loc[nm, d] = shift
                    if shift not in ["OFF", ""]:
                        work_counts[nm] += 1
                        if shift == "N": n_counts[nm] += 1
                        if shift == "DE": de_counts[nm] += 1
                        if roles[nm] == "HN" and shift == "E": hn_e_counts[nm] += 1

    # N 배치
    d = 1
    while d <= last_day:
        if "N" in df[d].values: d+=1; continue
        cands = [n for n in rn_list if n_counts[n] < MAX_N_LIMIT]
        cands.sort(key=lambda x: (n_counts[x], random.random())) 
        assigned = False
        for nm in cands:
            if roles[nm] == "HN": continue 
            if df.loc[nm, d] != "": continue
            allowed = fixed_work_map.get(nm, {}).get(d)
            rem = last_day - d + 1
            n_left = MAX_N_LIMIT - n_counts[nm]
            if n_left < 1: continue
            lengths = [3, 2] 
            if rem < 2: lengths = [1]
            random.shuffle(lengths)
            for length in lengths:
                if d + length - 1 > last_day: continue 
                if length > n_left: continue
                valid = True
                for i in range(length):
                    allow_i = fixed_work_map.get(nm, {}).get(d+i)
                    if not check_possibility(df, nm, d+i, "N", req_off_map.get(nm, []), allow_i, strict=True):
                        valid = False; break
                if valid and d+length <= last_day:
                    if df.loc[nm, d+length] not in ["", "OFF"]: valid = False
                if valid:
                    for i in range(length):
                        df.loc[nm, d+i] = "N"
                        n_counts[nm] += 1; work_counts[nm] += 1
                    if d+length <= last_day: 
                        if df.loc[nm, d+length] == "": df.loc[nm, d+length] = "OFF"
                    assigned = True; break
            if assigned: break
        
        # N 강제
        if not assigned and "N" not in df[d].values:
            for nm in rn_list:
                if roles[nm] == "HN": continue
                if n_counts[nm] >= MAX_N_LIMIT: continue 
                rem = last_day - d + 1
                if rem == 1:
                    if df.loc[nm, d] == "":
                         if check_possibility(df, nm, d, "N", req_off_map.get(nm, []), None, strict=True):
                            df.loc[nm, d] = "N"
                            n_counts[nm] += 1; work_counts[nm] += 1
                            assigned = True; break
                else:
                    if df.loc[nm, d] == "" and df.loc[nm, d+1] == "":
                         if n_counts[nm] + 2 > MAX_N_LIMIT: continue
                         p1 = check_possibility(df, nm, d, "N", req_off_map.get(nm, []), None, strict=True)
                         p2 = check_possibility(df, nm, d+1, "N", req_off_map.get(nm, []), None, strict=True)
                         if p1 and p2:
                            df.loc[nm, d] = "N"; df.loc[nm, d+1] = "N"
                            n_counts[nm] += 2; work_counts[nm] += 2
                            assigned = True; break
        d += 1

    # D, E, DE 배치
    for d in range(1, last_day + 1):
        is_hol = is_holiday_or_weekend(year, month, d, hol_set)
        needed = []
        if is_hol:
            if "DE" not in df[d].values: needed.append("DE")
        else:
            if "D" not in df[d].values: needed.append("D")
            if "E" not in df[d].values: needed.append("E")
            
        for shift in needed:
            candidates = [n for n in target_staff]
            def sort_key(nm):
                priority = 10
                if roles[nm] == "HN": 
                    if shift == "E": priority = 999 
                    else: priority = 0 
                return (priority, work_counts[nm], random.random())
            candidates.sort(key=sort_key)
            filled = False
            for p in candidates:
                if df.loc[p, d] != "": continue
                if shift == "DE" and de_counts[p] >= 1: continue
                if roles[p] == "HN" and is_hol and shift == "E" and hn_e_counts[p] >= 1: continue
                allowed = fixed_work_map.get(p, {}).get(d)
                if check_possibility(df, p, d, shift, req_off_map.get(p, []), allowed, strict=True):
                    df.loc[p, d] = shift
                    work_counts[p] += 1
                    if shift == "DE": de_counts[p] += 1
                    if roles[p] == "HN" and shift == "E": hn_e_counts[p] += 1
                    filled = True; break
            if not filled:
                for p in candidates:
                    if df.loc[p, d] != "": continue
                    if shift == "DE" and de_counts[p] >= 1: continue
                    if roles[p] == "HN" and is_hol and shift == "E" and hn_e_counts[p] >= 1: continue
                    if check_possibility(df, p, d, shift, req_off_map.get(p, []), None, strict=True):
                        df.loc[p, d] = shift
                        work_counts[p] += 1
                        if shift == "DE": de_counts[p] += 1
                        if roles[p] == "HN" and shift == "E": hn_e_counts[p] += 1
                        filled = True; break
            if not filled and shift != "DE":
                for p in candidates:
                    if df.loc[p, d] == "":
                        if check_possibility(df, p, d, shift, req_off_map.get(p, []), None, strict=True):
                            df.loc[p, d] = shift
                            work_counts[p] += 1
                            filled = True; break
            if not filled and shift == "DE": needed.extend(["D", "E"])

    # Zero Gap
    for d in range(1, last_day + 1):
        is_hol = is_holiday_or_weekend(year, month, d, hol_set)
        required = []
        current_shifts = list(df[d].values)
        if "N" not in current_shifts: required.append("N")
        if is_hol:
            if "DE" not in current_shifts and ("D" not in current_shifts or "E" not in current_shifts):
                if "D" not in current_shifts: required.append("D")
                if "E" not in current_shifts: required.append("E")
        else:
            if "D" not in current_shifts: required.append("D")
            if "E" not in current_shifts: required.append("E")
            
        for req_shift in required:
            limit_attempts = [MAX_N_LIMIT]
            if req_shift == "N": limit_attempts.append(MAX_N_EXTENDED)
            filled = False
            for n_limit in limit_attempts:
                if filled: break
                cands = [n for n in target_staff if df.loc[n, d] in ["", "OFF"]]
                if req_shift == "N": cands = [n for n in rn_list if roles[n] != "HN" and df.loc[n, d] in ["", "OFF"]]
                cands.sort(key=lambda x: (n_counts.get(x, 0), work_counts[x]))
                for p in cands:
                    if df.loc[p, d] in ["D", "E", "N", "DE"]: continue
                    if req_shift == "N" and n_counts[p] >= n_limit: continue
                    current_off_cnt = list(df.loc[p]).count("OFF") + list(df.loc[p]).count("")
                    if current_off_cnt <= MIN_OFF_LIMIT: continue 
                    if check_possibility(df, p, d, req_shift, req_off_map.get(p, []), None, strict=True):
                        df.loc[p, d] = req_shift
                        work_counts[p] += 1
                        if req_shift == "N": 
                            n_counts[p] += 1
                            if d < last_day and df.loc[p, d+1] in ["", "OFF"]:
                                 if n_counts[p] < n_limit and (current_off_cnt - 2) >= MIN_OFF_LIMIT:
                                     if check_possibility(df, p, d+1, "N", req_off_map.get(p, []), None, strict=True):
                                         df.loc[p, d+1] = "N"
                                         n_counts[p] += 1; work_counts[p] += 1
                        filled = True; break

    # Equalizer
    for _ in range(50):
        temp_offs = {n: 0 for n in rn_list}
        for n in rn_list:
            cnt = list(df.loc[n]).count("OFF") + list(df.loc[n]).count("")
            weight = 0.3
            if n_counts[n] > MAX_N_LIMIT: weight = 1.0 
            adjusted = cnt - (n_counts[n] * weight)
            temp_offs[n] = adjusted
        if not temp_offs: break
        max_p = max(temp_offs, key=temp_offs.get)
        min_p = min(temp_offs, key=temp_offs.get)
        if temp_offs[max_p] - temp_offs[min_p] <= 2: break
        days = list(range(1, last_day+1)); random.shuffle(days)
        for d in days:
            if df.loc[max_p, d] not in ["", "OFF"]: continue
            task = df.loc[min_p, d]
            if task not in ["D", "E"]: continue
            if check_possibility(df, max_p, d, task, req_off_map.get(max_p, []), strict=True):
                df.loc[max_p, d] = task
                df.loc[min_p, d] = "OFF"
                work_counts[max_p] += 1; work_counts[min_p] -= 1
                break

    # 최소 OFF
    for nm in target_staff:
        current_off = list(df.loc[nm]).count("OFF") + list(df.loc[nm]).count("")
        if current_off < MIN_OFF_LIMIT:
            needed = MIN_OFF_LIMIT - current_off
            days = list(range(1, last_day+1)); random.shuffle(days)
            for d in days:
                if needed <= 0: break
                if df.loc[nm, d] in ["D", "E"]: 
                    day_shifts = list(df[d].values)
                    shift = df.loc[nm, d]
                    if day_shifts.count(shift) > 1:
                        df.loc[nm, d] = "OFF"
                        work_counts[nm] -= 1
                        needed -= 1
        current_off = list(df.loc[nm]).count("OFF") + list(df.loc[nm]).count("")
        while current_off < MIN_OFF_LIMIT:
            candidates_days = [d for d in range(1, last_day+1) if df.loc[nm, d] in ["D", "E"]]
            if not candidates_days: break
            random.shuffle(candidates_days)
            target_d = candidates_days[0]
            target_shift = df.loc[nm, target_d]
            df.loc[nm, target_d] = "OFF"
            current_off += 1
            work_counts[nm] -= 1
            replacements = [r for r in rn_list if df.loc[r, target_d] in ["", "OFF"]]
            replacements.sort(key=lambda x: -(list(df.loc[x]).count("OFF") + list(df.loc[x]).count("")))
            for r in replacements:
                r_off = list(df.loc[r]).count("OFF") + list(df.loc[r]).count("")
                if r_off > MIN_OFF_LIMIT:
                    if check_possibility(df, r, target_d, target_shift, req_off_map.get(r, []), strict=True):
                        df.loc[r, target_d] = target_shift
                        work_counts[r] += 1
                        break 

    df.fillna("OFF", inplace=True)
    df.replace("", "OFF", inplace=True)
    return True, df, req_off_map, n_counts

# ==========================================
# 5. 시뮬레이션
# ==========================================
def run_simulation(year, month, staff_data):
    last_day = calendar.monthrange(year, month)[1]
    hol_set = get_holidays_in_month(year, month)
    best_df = None; best_req_map = None; min_score = 999999
    
    for i in range(100): 
        success, df, req_map, n_cnts = attempt_schedule(year, month, staff_data, hol_set, last_day)
        
        has_hole = False
        for d in range(1, last_day+1):
            day_shifts = list(df[d].values)
            if "N" not in day_shifts: has_hole = True
            is_hol = is_holiday_or_weekend(year, month, d, hol_set)
            if is_hol:
                if "DE" not in day_shifts and ("D" not in day_shifts or "E" not in day_shifts): has_hole = True
            else:
                if "D" not in day_shifts: has_hole = True
                if "E" not in day_shifts: has_hole = True
        
        if success:
            rn_offs = []
            single_offs = 0
            long_offs = 0 
            min_off_violation = 0
            max_n_violation = 0
            
            for n in df.index:
                row = list(df.loc[n])
                role = next((s['role'] for s in staff_data if s['name']==n), "")
                off_cnt = row.count("OFF")
                n_c = row.count("N")
                
                if role == "RN": rn_offs.append(off_cnt)
                if off_cnt < MIN_OFF_LIMIT: min_off_violation += 1
                if n_c > MAX_N_EXTENDED: max_n_violation += 1 
                
                cons_off = 0
                for idx in range(len(row)):
                    if row[idx] == "OFF":
                        cons_off += 1
                    else:
                        if cons_off >= 4: long_offs += 1
                        cons_off = 0
                        
                for idx in range(1, len(row)-1):
                    if row[idx] == "OFF" and row[idx-1] != "OFF" and row[idx+1] != "OFF":
                        single_offs += 1
            
            diff = max(rn_offs) - min(rn_offs) if rn_offs else 0
            score = (diff * 50) + (single_offs * 30) + (long_offs * 40)
            if has_hole: score += 9999999
            score += (min_off_violation * 999999)
            score += (max_n_violation * 999999)
            
            if best_df is None or score < min_score:
                min_score = score
                best_df = df
                best_req_map = req_map
            
            if not has_hole and min_off_violation == 0 and max_n_violation == 0 and diff <= 2 and single_offs <= 3 and long_offs == 0: break
                
    if best_df is None:
        names = [s['name'] for s in staff_data]
        best_df = pd.DataFrame("OFF", index=names, columns=range(1, last_day + 1))

    return best_df, best_req_map

# ★ 엑셀 파일 파싱 함수 (업로드용)
def parse_uploaded_excel(uploaded_file):
    try:
        # 헤더 없이 일단 읽음
        raw_df = pd.read_excel(uploaded_file, header=None)
        
        # '이름'이 있는 행 찾기 (헤더 행)
        header_row_idx = None
        for i, row in raw_df.iterrows():
            if "이름" in row.values:
                header_row_idx = i
                break
        
        if header_row_idx is None: return None
        
        # 헤더 적용하여 다시 생성
        df = raw_df.iloc[header_row_idx+1:].copy()
        df.columns = raw_df.iloc[header_row_idx].values
        
        # 이름 컬럼을 인덱스로 설정
        if "이름" not in df.columns: return None
        df.set_index("이름", inplace=True)
        
        # 날짜 컬럼만 추출 (1, 2, 3... 31)
        valid_cols = [c for c in df.columns if isinstance(c, int)]
        clean_df = df[valid_cols].fillna("OFF").astype(str)
        
        # 요일, 합계 행 등 제거 (이름이 실제 직원 명단에 있는 경우만 필터링)
        # staff_db에 있는 이름만 남김
        current_staff = [s['name'] for s in st.session_state.staff_list]
        clean_df = clean_df[clean_df.index.isin(current_staff)]
        
        return clean_df
    except:
        return None

def prepare_display_df(df, year, month, staff_data):
    if df is None: return None
    last_day = df.shape[1]
    hol_set = get_holidays_in_month(year, month)
    legal_off = sum(1 for d in range(1, last_day+1) if is_holiday_or_weekend(year, month, d, hol_set))
    
    disp = df.copy().fillna("OFF").astype(str)
    no_l, name_l, role_l, n_l, off_l, ann_l, un_l = [], [], [], [], [], [], []
    count = 1
    staff_info = {s['name']: s for s in staff_data}
    
    for nm in disp.index:
        row = list(disp.loc[nm])
        n_c = row.count("N")
        off_c = row.count("OFF")
        info = staff_info.get(nm, {'role': '', 'annual_leave': 0})
        no_l.append(str(count)); name_l.append(nm); role_l.append(info['role'])
        n_l.append(str(n_c)); off_l.append(str(off_c))
        ann_l.append(str(info['annual_leave'])); un_l.append(str(max(0, legal_off - off_c)))
        count += 1
        
    disp.reset_index(drop=True, inplace=True)
    disp.insert(0, "직군", role_l); disp.insert(0, "이름", name_l); disp.insert(0, "No.", no_l)
    disp["N"] = n_l; disp["OFF"] = off_l; disp["연차"] = ann_l; disp["미사용OFF"] = un_l
    
    summary = {str(c): [] for c in disp.columns}
    for task in ["D", "E", "N", "DE", "M"]:
        for col in disp.columns:
            if str(col) == "이름": summary[str(col)].append(task)
            elif str(col) in ["No.", "직군", "N", "OFF", "연차", "미사용OFF"]: summary[str(col)].append("")
            else:
                try: cnt = list(disp[col]).count(task); summary[str(col)].append(str(cnt))
                except: summary[str(col)].append("")

    sum_df = pd.DataFrame(summary, index=["D", "E", "N", "DE", "M"])
    disp.columns = disp.columns.astype(str)
    final = pd.concat([disp, sum_df])
    return final

def apply_browser_style(v):
    base = "background-color: #FFFFFF; color: #000000; text-align: center; border: 1px solid #f0f0f0;"
    if v == "N": return f"background-color: #{COLOR_N}; color: #000000; text-align: center;"
    elif v == "DE": return f"background-color: #{COLOR_DE}; color: #000000; text-align: center;"
    return base

# ==========================================
# 7. 엑셀 출력
# ==========================================
def to_excel(df, year, month, req_off_map, staff_data):
    clean_df = df.copy()
    new_cols = []
    for c in clean_df.columns:
        if str(c).isdigit(): new_cols.append(int(c))
        else: new_cols.append(c)
    clean_df.columns = new_cols
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}월"
    
    align_c = Alignment(horizontal='center', vertical='center')
    align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_title = Font(name='맑은 고딕', size=20, bold=True)
    font_bold = Font(name='맑은 고딕', size=11, bold=True)
    font_norm = Font(name='맑은 고딕', size=11, bold=False)
    
    fill_wk = PatternFill('solid', fgColor=COLOR_WEEKEND_BG)
    fill_n = PatternFill('solid', fgColor=COLOR_N)
    fill_de = PatternFill('solid', fgColor=COLOR_DE)
    fill_req = PatternFill('solid', fgColor=COLOR_REQ_OFF)
    fill_err = PatternFill('solid', fgColor=COLOR_ERROR)
    
    day_cols = [c for c in clean_df.columns if isinstance(c, int)]
    last_day = max(day_cols) if day_cols else 28
    hol_set = get_holidays_in_month(year, month)
    legal_off = sum(1 for d in range(1, last_day+1) if is_holiday_or_weekend(year, month, d, hol_set))
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_day+6)
    ws['A1'] = f"{year}년 {month}월 근무표 (OFF: {legal_off} )"
    ws['A1'].font = font_title; ws['A1'].alignment = align_c
    ws.row_dimensions[1].height = 50
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws['A2'] = "부서 : 5병동"
    ws['A2'].font = Font(name='맑은 고딕', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    header_row = 3; day_row = 4; data_start_row = 5
    
    ws.cell(row=header_row, column=1, value="직군").font = font_bold
    ws.cell(row=header_row, column=1).alignment = align_c; ws.cell(row=header_row, column=1).border = border
    ws.cell(row=header_row, column=2, value="이름").font = font_bold
    ws.cell(row=header_row, column=2).alignment = align_c; ws.cell(row=header_row, column=2).border = border
    
    ws.cell(row=day_row, column=1).border = border; ws.cell(row=day_row, column=2).border = border
    
    weekend_cols = set()
    days_str = ["월", "화", "수", "목", "금", "토", "일"]
    
    for d in range(1, last_day+1):
        dt = date(year, month, d)
        is_wk = is_holiday_or_weekend(year, month, d, hol_set)
        col_idx = d + 2
        if is_wk: weekend_cols.add(col_idx)
        c2 = ws.cell(row=header_row, column=col_idx, value=d)
        c2.font = font_bold; c2.alignment = align_c; c2.border = border
        c3 = ws.cell(row=day_row, column=col_idx, value=days_str[dt.weekday()])
        c3.font = font_bold; c3.alignment = align_c; c3.border = border
        if is_wk: c2.fill = fill_wk; c3.fill = fill_wk
        ws.column_dimensions[get_column_letter(col_idx)].width = 3.8

    stats_h = ["N", "OFF", "연차", "미사용\nOFF"]
    for i, t in enumerate(stats_h):
        col_idx = last_day + 3 + i
        c = ws.cell(row=day_row, column=col_idx, value=t)
        c.font = font_bold; c.alignment = align_wrap; c.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 8.0

    current_row = data_start_row
    groups = ["HN", "RN", "AN"]
    staff_records = staff_data
    ws.column_dimensions['A'].width = 6.0
    ws.column_dimensions['B'].width = 11.0
    duty_codes = ["D", "E", "N", "DE", "M"]
    
    for i, grp in enumerate(groups):
        mems = [nm for nm in clean_df.index if next((s['role'] for s in staff_records if s['name']==nm), "")==grp]
        for nm in mems:
            ws.row_dimensions[current_row].height = 32
            ws.cell(row=current_row, column=1, value=grp).font = font_bold
            ws.cell(row=current_row, column=1).alignment = align_c
            ws.cell(row=current_row, column=1).border = border
            
            ws.cell(row=current_row, column=2, value=nm).font = font_bold
            ws.cell(row=current_row, column=2).alignment = align_c
            ws.cell(row=current_row, column=2).border = border
            
            n_c, off_c = 0, 0
            for d in range(1, last_day+1):
                try: val = clean_df.loc[nm, d]
                except: val = "OFF"
                col_idx = d + 2
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = align_c; cell.border = border
                if val in duty_codes: cell.font = font_norm
                else: cell.font = font_bold
                if col_idx in weekend_cols: cell.fill = fill_wk
                if val=="N": n_c+=1; cell.fill = fill_n
                if val=="DE": cell.fill = fill_de
                safe_map = req_off_map if req_off_map is not None else {}
                if nm in safe_map and d in safe_map[nm] and val=="OFF": cell.fill = fill_req
                if val=="OFF": off_c+=1
            
            s_info = next((s for s in staff_records if s['name'] == nm), None)
            ann = s_info['annual_leave'] if s_info else 0
            unused = max(0, legal_off - off_c)
            vals = [n_c, off_c, ann, unused]
            for idx, v in enumerate(vals):
                c = ws.cell(row=current_row, column=last_day+3+idx, value=v)
                c.font = font_bold; c.alignment = align_c; c.border = border
            current_row += 1
        if grp == "RN" and i < len(groups)-1: current_row += 1
    
    current_row += 1
    for job in duty_codes:
        ws.row_dimensions[current_row].height = 32
        ws.cell(row=current_row, column=1, value="").border = border
        ws.cell(row=current_row, column=2, value=job).font = font_bold
        ws.cell(row=current_row, column=2).alignment = align_c
        ws.cell(row=current_row, column=2).border = border
        
        for d in range(1, last_day+1):
            col_vals = list(clean_df[d])
            cnt = col_vals.count(job)
            col_idx = d + 2
            cell = ws.cell(row=current_row, column=col_idx, value=cnt)
            cell.font = font_bold; cell.alignment = align_c; cell.border = border
            if col_idx in weekend_cols: cell.fill = fill_wk
            
            is_err = False
            if job == "N" and cnt < 1: is_err = True
            is_hol = is_holiday_or_weekend(year, month, d, hol_set)
            if not is_hol and job in ["D", "E"] and cnt < 1: is_err = True
            if is_err: cell.fill = fill_err

        current_row += 1
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

# ==========================================
# 8. UI
# ==========================================
st.title("5병동 근무표 관리")

with st.sidebar:
    st.header("설정 및 명단")
    col1, col2 = st.columns(2)
    s_year = col1.number_input("년도", 2026, 2027, 2026)
    s_month = col2.selectbox("월", range(1, 13), index=1)
    
    saved_files = glob.glob(f"{SCHEDULE_DIR}/schedule_*.csv")
    saved_files.sort(reverse=True)
    
    st.divider()
    st.subheader("📂 저장된 근무표 목록")
    
    # 엑셀 업로드 추가
    uploaded_file = st.file_uploader("📤 엑셀 파일 업로드 (기존 파일 덮어쓰기)", type=['xlsx'])
    if uploaded_file:
        parsed_df = parse_uploaded_excel(uploaded_file)
        if parsed_df is not None:
            # 파일명을 기반으로 저장하거나, 현재 선택된 월로 저장
            st.session_state.df_res = parsed_df
            st.session_state.view_mode = "history"
            st.session_state.history_ym = (s_year, s_month) # 현재 선택된 월로 가정
            st.success("엑셀 파일이 로드되었습니다. 확인 후 '확정' 버튼을 누르면 저장됩니다.")
        else:
            st.error("엑셀 파일 형식이 올바르지 않습니다.")

    for f in saved_files:
        fname = os.path.basename(f)
        parts = fname.replace("schedule_", "").replace(".csv", "").split("_")
        if len(parts) == 2:
            y, m = parts
            c1, c2 = st.columns([4, 1])
            if c1.button(f"{y}년 {m}월 보기", key=f"btn_{y}_{m}", use_container_width=True):
                loaded_df = load_schedule_file_path(f)
                if loaded_df is not None:
                    st.session_state.df_res = loaded_df
                    st.session_state.req_map = {}
                    st.session_state.view_mode = "history"
                    st.session_state.history_ym = (int(y), int(m))
                    st.rerun()
            if c2.button("🗑️", key=f"del_{y}_{m}"):
                delete_schedule_file(int(y), int(m))
                st.rerun()

    st.divider()
    
    if st.button("🔄 Request Off 일괄 초기화", type="secondary", use_container_width=True):
        st.session_state.backup_staff_list = [s.copy() for s in st.session_state.staff_list]
        for s in st.session_state.staff_list: 
            s['req_off'] = ""
            s['fixed_work'] = ""
        save_staff_data(st.session_state.staff_list); st.rerun()
    
    if st.button("🔄 고정 근무만 초기화", type="secondary", use_container_width=True):
        for s in st.session_state.staff_list:
             s['fixed_work'] = ""
        save_staff_data(st.session_state.staff_list); st.rerun()

    if st.session_state.backup_staff_list:
        if st.button("↩️ 실행 취소 (되돌리기)", type="primary", use_container_width=True):
            st.session_state.staff_list = st.session_state.backup_staff_list
            st.session_state.backup_staff_list = None
            save_staff_data(st.session_state.staff_list); st.rerun()
    st.divider()
    
    if st.session_state.edit_index is None:
        st.write("**👥 근무자 목록 (RN 고연차순)**")
        with st.form("add"):
            c1,c2 = st.columns(2)
            name = c1.text_input("이름")
            role = c2.selectbox("직군", ROLES)
            off = st.text_input("Request Off")
            fixed = st.text_input("고정근무 (예: 1=N/OFF)")
            annual = st.number_input("연차",0,30,0)
            if st.form_submit_button("추가"):
                st.session_state.staff_list.append({"name":name,"role":role,"req_off":off,"annual_leave":annual, "fixed_work":fixed})
                save_staff_data(st.session_state.staff_list); st.rerun()
        
        if st.session_state.staff_list:
            st.write("---")
            for i, s in enumerate(st.session_state.staff_list):
                c1,c2,c3,c4,c5 = st.columns([3,2,1,1,1])
                with c1: st.write(f"**{i+1}. {s['name']}** ({s['role']})")
                if s.get('req_off'): c1.caption(f"Request Off: {s['req_off']}")
                if s.get('fixed_work'): c1.caption(f"고정: {s['fixed_work']}")
                if c2.button("수정", key=f"e{i}"): st.session_state.edit_index = i; st.rerun()
                if c3.button("▲", key=f"u{i}") and i>0:
                    st.session_state.staff_list[i], st.session_state.staff_list[i-1] = st.session_state.staff_list[i-1], st.session_state.staff_list[i]
                    save_staff_data(st.session_state.staff_list); st.rerun()
                if c4.button("▼", key=f"d{i}") and i<len(st.session_state.staff_list)-1:
                    st.session_state.staff_list[i], st.session_state.staff_list[i+1] = st.session_state.staff_list[i+1], st.session_state.staff_list[i]
                    save_staff_data(st.session_state.staff_list); st.rerun()
                if c5.button("❌", key=f"x{i}"):
                    st.session_state.staff_list.pop(i); save_staff_data(st.session_state.staff_list); st.rerun()
    else:
        idx = st.session_state.edit_index
        t = st.session_state.staff_list[idx]
        st.info(f"✏️ **{t['name']}** 수정")
        with st.form("edit"):
            nn = st.text_input("이름", t['name'])
            nr = st.selectbox("직군", ROLES, ROLES.index(t['role']))
            no = st.text_input("Request Off (예: 1, 5, 10)", str(t.get('req_off','')))
            nf = st.text_input("고정근무 (예: 1=N/OFF, 5=D)", str(t.get('fixed_work','')), placeholder="예: 1=N/OFF, 5=D")
            na = st.number_input("연차", 0, 30, int(t.get('annual_leave',0)))
            if st.form_submit_button("저장"):
                st.session_state.staff_list[idx] = {"name":nn,"role":nr,"req_off":no,"annual_leave":na, "fixed_work":nf}
                save_staff_data(st.session_state.staff_list); st.session_state.edit_index=None; st.rerun()
            if st.form_submit_button("취소"): st.session_state.edit_index=None; st.rerun()

# 메인 화면
if st.session_state.view_mode == "history" and st.session_state.df_res is not None:
    y, m = st.session_state.history_ym
    st.info(f"📂 불러온 근무표: {y}년 {m}월")
    
    # 직접 수정 모드
    if st.checkbox("✏️ 근무표 직접 수정하기", value=st.session_state.edit_mode):
        st.session_state.edit_mode = True
        
        # 전체 데이터프레임(통계 포함) 생성
        full_df = prepare_display_df(st.session_state.df_res, y, m, st.session_state.staff_list)
        # 에디터로 표시 (전체 수정 가능)
        edited_full_df = st.data_editor(full_df, use_container_width=True, height=600)
        
        if st.button("💾 수정사항 저장"):
            # 저장 시에는 통계/합계 행을 제외하고 순수 근무표 데이터만 추출해야 함
            # 인덱스가 직원 이름인 행만 필터링, 컬럼은 1~말일만
            staff_names = [s['name'] for s in st.session_state.staff_list]
            # 인덱스 리셋된 상태이므로 '이름' 컬럼 사용
            valid_rows = edited_full_df[edited_full_df['이름'].isin(staff_names)].copy()
            valid_rows.set_index('이름', inplace=True)
            
            # 날짜 컬럼만 추출 (1, 2, 3...)
            date_cols = [c for c in valid_rows.columns if str(c).isdigit()]
            final_df = valid_rows[date_cols]
            
            st.session_state.df_res = final_df
            save_schedule_file(final_df, y, m)
            st.success("저장되었습니다!")
            time.sleep(1)
            st.rerun()
            
    else:
        st.session_state.edit_mode = False
        display_df = prepare_display_df(st.session_state.df_res, y, m, st.session_state.staff_list)
        st.dataframe(display_df.style.map(apply_browser_style), use_container_width=True, height=600, hide_index=True)
    
    c1, c2 = st.columns([1, 1])
    with c1:
        xlsx = to_excel(st.session_state.df_res, y, m, st.session_state.req_map, st.session_state.staff_list)
        st.download_button("📄 엑셀 다운로드", xlsx, f"5병동_{y}_{m}_확정.xlsx", use_container_width=True)
    with c2:
        if st.button("돌아가기 (새 생성)", use_container_width=True):
            st.session_state.view_mode = "create"
            del st.session_state.df_res
            st.rerun()

else:
    # 생성 모드
    if st.button("🎲 근무표 생성", type="primary", use_container_width=True):
        if not st.session_state.staff_list: st.error("근무자 없음")
        else:
            with st.spinner("생성 중..."):
                best_df, best_req_map = run_simulation(s_year, s_month, st.session_state.staff_list)
            
            if best_df is not None:
                st.session_state.df_res = best_df
                st.session_state.req_map = best_req_map
                st.session_state.prev_df_res = best_df
                st.session_state.prev_req_map = best_req_map
                st.rerun()
            else: 
                st.error("생성 실패")

    if "df_res" in st.session_state and st.session_state.df_res is not None:
        st.divider()
        st.markdown("### 📝 생성 결과 (미확정)")
        
        # 직접 수정 모드 (생성 직후)
        if st.checkbox("✏️ 근무표 직접 수정하기", value=st.session_state.edit_mode):
            st.session_state.edit_mode = True
            
            full_df = prepare_display_df(st.session_state.df_res, s_year, s_month, st.session_state.staff_list)
            edited_full_df = st.data_editor(full_df, use_container_width=True, height=600)
            
            if st.button("💾 수정사항 반영"):
                staff_names = [s['name'] for s in st.session_state.staff_list]
                valid_rows = edited_full_df[edited_full_df['이름'].isin(staff_names)].copy()
                valid_rows.set_index('이름', inplace=True)
                date_cols = [c for c in valid_rows.columns if str(c).isdigit()]
                final_df = valid_rows[date_cols]
                
                st.session_state.df_res = final_df
                st.rerun()
        else:
            st.session_state.edit_mode = False
            display_df = prepare_display_df(st.session_state.df_res, s_year, s_month, st.session_state.staff_list)
            st.dataframe(display_df.style.map(apply_browser_style), use_container_width=True, height=600, hide_index=True)
        
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            xlsx = to_excel(st.session_state.df_res, s_year, s_month, st.session_state.req_map, st.session_state.staff_list)
            st.download_button("📄 엑셀 미리보기", xlsx, f"5병동_{s_year}_{s_month}_임시.xlsx", use_container_width=True)
        
        with c2:
            if st.button("✅ 확정 (저장)", type="primary", use_container_width=True):
                save_schedule_file(st.session_state.df_res, s_year, s_month)
                del st.session_state.df_res
                st.rerun()
        
        with c3:
            if st.button("🎲 재배정", use_container_width=True):
                 with st.spinner("재배정 중..."):
                    best_df, best_req_map = run_simulation(s_year, s_month, st.session_state.staff_list)
                    st.session_state.prev_df_res = st.session_state.df_res
                    st.session_state.df_res = best_df
                    st.session_state.req_map = best_req_map
                    st.rerun()

        if "prev_df_res" in st.session_state and st.session_state.prev_df_res is not None:
             if st.button("↩️ 실행 취소 (이전 결과 불러오기)"):
                  st.session_state.df_res = st.session_state.prev_df_res
                  st.rerun()